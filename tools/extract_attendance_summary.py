from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from datetime import date, time
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


TIME_RE = re.compile(r"(\d{1,2}):(\d{2})")
HEADER_DAY_RE = re.compile(r"^(\d{1,2})")
FILENAME_PERIOD_RE = re.compile(r"(?P<start>20\d{6})-(?P<end>20\d{6})")


@dataclass(frozen=True)
class Config:
    source: Path
    output: Path
    sheet_name: str
    year: int
    month: int
    overtime_start: time
    holidays: dict[date, str]


def parse_hhmm(value: str) -> time:
    match = re.fullmatch(r"(\d{1,2}):(\d{2})", value.strip())
    if not match:
        raise argparse.ArgumentTypeError("时间格式应为 HH:MM")
    hour = int(match.group(1))
    minute = int(match.group(2))
    if not (0 <= hour <= 23 and 0 <= minute <= 59):
        raise argparse.ArgumentTypeError("时间超出范围")
    return time(hour, minute)


def parse_holidays(values: list[str]) -> dict[date, str]:
    result: dict[date, str] = {}
    for item in values:
        day_text, _, name = item.partition(":")
        try:
          holiday = date.fromisoformat(day_text)
        except ValueError as exc:
          raise argparse.ArgumentTypeError(f"无效节假日日期: {item}") from exc
        result[holiday] = name or "法定节假日"
    return result


def infer_year_month(source: Path) -> tuple[int, int]:
    match = FILENAME_PERIOD_RE.search(source.name)
    if not match:
        today = date.today()
        return today.year, today.month
    start_text = match.group("start")
    return int(start_text[:4]), int(start_text[4:6])


def default_output(source: Path) -> Path:
    match = FILENAME_PERIOD_RE.search(source.name)
    suffix = match.group(0) if match else date.today().strftime("%Y%m%d")
    return Path("outputs") / f"员工出勤与加班汇总_{suffix}.xlsx"


def parse_args() -> Config:
    parser = argparse.ArgumentParser(description="员工出勤与加班统计工具")
    parser.add_argument("source", type=Path, help="打卡记录 Excel 文件路径")
    parser.add_argument("-o", "--output", type=Path, help="输出 Excel 文件路径")
    parser.add_argument("--sheet", default="打卡时间记录", help="源数据工作表名称")
    parser.add_argument("--year", type=int, help="统计年份，默认从文件名推断")
    parser.add_argument("--month", type=int, help="统计月份，默认从文件名推断")
    parser.add_argument("--overtime-start", type=parse_hhmm, default=time(18, 0), help="加班起算时间，格式 HH:MM")
    parser.add_argument("--holiday", action="append", default=[], help="法定节假日，格式 YYYY-MM-DD[:名称]，可重复传入")
    args = parser.parse_args()

    source = args.source.expanduser().resolve()
    if not source.exists():
        parser.error(f"源文件不存在: {source}")

    inferred_year, inferred_month = infer_year_month(source)
    return Config(
        source=source,
        output=(args.output or default_output(source)).expanduser().resolve(),
        sheet_name=args.sheet,
        year=args.year or inferred_year,
        month=args.month or inferred_month,
        overtime_start=args.overtime_start,
        holidays=parse_holidays(args.holiday),
    )


def parse_times(value: object) -> list[time]:
    if value is None:
        return []
    text = str(value).strip()
    if not text or text == "--":
        return []

    result: list[time] = []
    for hour_text, minute_text in TIME_RE.findall(text):
        hour = int(hour_text)
        minute = int(minute_text)
        if 0 <= hour <= 23 and 0 <= minute <= 59:
            result.append(time(hour, minute))
    return sorted(result)


def minutes_since_midnight(value: time) -> int:
    return value.hour * 60 + value.minute


def raw_overtime_minutes(times: list[time], overtime_start: time) -> int:
    if not times:
        return 0
    return max(0, minutes_since_midnight(times[-1]) - minutes_since_midnight(overtime_start))


def overtime_minutes(times: list[time], overtime_start: time) -> int:
    return (raw_overtime_minutes(times, overtime_start) // 30) * 30


def has_day_shift_pattern(times: list[time]) -> bool:
    return any(time(10, 0) <= item <= time(14, 30) for item in times)


def is_night_start(times: list[time]) -> bool:
    return bool(times) and any(item >= time(18, 0) for item in times) and not has_day_shift_pattern(times)


def first_morning_time(times: list[time]) -> time | None:
    morning_times = [item for item in times if item <= time(8, 59)]
    return min(morning_times) if morning_times else None


def night_shift_type(end_time: time | None) -> str:
    if end_time is None:
        return "未闭合"
    if end_time >= time(8, 0):
        return "大夜班"
    return "小夜班"


def parse_date(header: object, year: int, month: int) -> date:
    match = HEADER_DAY_RE.match(str(header or "").strip())
    if not match:
        raise ValueError(f"无法识别日期表头: {header!r}")
    return date(year, month, int(match.group(1)))


def day_type(work_date: date, holidays: dict[date, str]) -> str:
    if work_date in holidays:
        return "法定节假日"
    if work_date.weekday() == 5:
        return "周六"
    if work_date.weekday() == 6:
        return "周日"
    return "工作日"


def auto_width(sheet) -> None:
    for col_idx in range(1, sheet.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = max((len(str(cell.value)) for cell in sheet[letter] if cell.value is not None), default=0)
        sheet.column_dimensions[letter].width = min(max(max_len + 2, 10), 48)


def style_table(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    auto_width(sheet)


def load_source(config: Config):
    try:
        wb = openpyxl.load_workbook(config.source, data_only=True)
    except Exception as exc:
        raise RuntimeError(f"无法读取 Excel 文件: {config.source}") from exc
    if config.sheet_name not in wb.sheetnames:
        raise RuntimeError(f"找不到工作表 {config.sheet_name!r}，可用工作表: {', '.join(wb.sheetnames)}")
    return wb[config.sheet_name]


def build_report(config: Config) -> tuple[openpyxl.Workbook, int, int]:
    ws = load_source(config)
    date_headers = [
        (ws.cell(row=4, column=col).value, parse_date(ws.cell(row=4, column=col).value, config.year, config.month))
        for col in range(6, ws.max_column + 1)
    ]
    rows = []
    detail_rows = []
    date_type_rows = []

    for header, work_date in date_headers:
        date_type_rows.append([work_date.isoformat(), str(header).replace("\n", " "), day_type(work_date, config.holidays), config.holidays.get(work_date, "")])

    for row_idx in range(5, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=1).value
        account = ws.cell(row=row_idx, column=2).value
        department = ws.cell(row=row_idx, column=3).value
        job_title = ws.cell(row=row_idx, column=4).value
        employee_id = ws.cell(row=row_idx, column=5).value
        if not name:
            continue

        summary = {
            "attendance_days": 0,
            "workday_attendance": 0,
            "holiday_attendance": 0,
            "saturday_attendance": 0,
            "sunday_attendance": 0,
            "total_overtime": 0,
            "workday_overtime": 0,
            "weekend_overtime": 0,
            "holiday_overtime": 0,
            "overtime_days": 0,
            "small_night_count": 0,
            "big_night_count": 0,
            "unclosed_night_count": 0,
        }

        day_records = []
        for offset, (header, work_date) in enumerate(date_headers, start=6):
            raw_value = ws.cell(row=row_idx, column=offset).value
            day_records.append({"header": header, "work_date": work_date, "raw_value": raw_value, "times": parse_times(raw_value)})

        for day_index, record in enumerate(day_records):
            work_date = record["work_date"]
            raw_value = record["raw_value"]
            times = record["times"]
            current_day_type = day_type(work_date, config.holidays)
            night_start = is_night_start(times)
            next_morning = None
            current_night_type = ""

            if night_start:
                next_record = day_records[day_index + 1] if day_index + 1 < len(day_records) else None
                next_morning = first_morning_time(next_record["times"]) if next_record else None
                current_night_type = night_shift_type(next_morning)
                if current_night_type == "小夜班":
                    summary["small_night_count"] += 1
                elif current_night_type == "大夜班":
                    summary["big_night_count"] += 1
                else:
                    summary["unclosed_night_count"] += 1

            if times:
                summary["attendance_days"] += 1
                if current_day_type == "法定节假日":
                    summary["holiday_attendance"] += 1
                elif current_day_type == "周六":
                    summary["saturday_attendance"] += 1
                elif current_day_type == "周日":
                    summary["sunday_attendance"] += 1
                else:
                    summary["workday_attendance"] += 1

            minutes = 0 if night_start else overtime_minutes(times, config.overtime_start)
            raw_minutes = 0 if night_start else raw_overtime_minutes(times, config.overtime_start)
            if minutes:
                summary["overtime_days"] += 1
                summary["total_overtime"] += minutes
                if current_day_type == "法定节假日":
                    summary["holiday_overtime"] += minutes
                elif current_day_type in ("周六", "周日"):
                    summary["weekend_overtime"] += minutes
                else:
                    summary["workday_overtime"] += minutes

            detail_rows.append([
                name, account, employee_id, work_date.isoformat(), record["header"], current_day_type, raw_value,
                1 if times else 0, "是" if night_start else "否", current_night_type,
                next_morning.strftime("%H:%M") if next_morning else "", round(raw_minutes / 60, 2), raw_minutes,
                round(minutes / 60, 2), minutes,
                round(minutes / 60, 2) if current_day_type == "工作日" else 0,
                round(minutes / 60, 2) if current_day_type in ("周六", "周日") else 0,
                round(minutes / 60, 2) if current_day_type == "法定节假日" else 0,
            ])

        rows.append([
            name, account, department, job_title, employee_id,
            summary["attendance_days"], summary["workday_attendance"], summary["holiday_attendance"],
            summary["saturday_attendance"], summary["sunday_attendance"],
            round(summary["total_overtime"] / 60, 2), summary["total_overtime"],
            round(summary["workday_overtime"] / 60, 2), summary["workday_overtime"],
            round(summary["weekend_overtime"] / 60, 2), summary["weekend_overtime"],
            round(summary["holiday_overtime"] / 60, 2), summary["holiday_overtime"],
            summary["small_night_count"] + summary["big_night_count"], summary["small_night_count"],
            summary["big_night_count"], summary["unclosed_night_count"], summary["overtime_days"],
        ])

    out = openpyxl.Workbook()
    summary_sheet = out.active
    summary_sheet.title = "出勤加班汇总"
    summary_sheet.append(["姓名", "账号", "部门", "职务", "工号", "出勤天数", "工作日出勤天数", "法定节假日出勤天数", "周六出勤天数", "周日出勤天数", "加班时长(小时)", "加班时长(分钟)", "平时加班时长(小时)", "平时加班时长(分钟)", "周末加班时长(小时)", "周末加班时长(分钟)", "法定假期加班时长(小时)", "法定假期加班时长(分钟)", "夜班次数", "小夜班次数", "大夜班次数", "未闭合夜班次数", "加班天数"])
    for item in rows:
        summary_sheet.append(item)
    style_table(summary_sheet)

    detail = out.create_sheet("每日明细")
    detail.append(["姓名", "账号", "工号", "日期", "原表日期", "日期类型", "原始打卡记录", "是否出勤", "是否夜班开始", "夜班类型", "次日早晨下班时间", "原始加班小时", "原始加班分钟", "规则后加班小时", "规则后加班分钟", "当日平时加班小时", "当日周末加班小时", "当日法定假期加班小时"])
    for item in detail_rows:
        detail.append(item)
    style_table(detail)

    date_sheet = out.create_sheet("日期分类")
    date_sheet.append(["日期", "原表日期", "日期类型", "法定节假日名称"])
    for item in date_type_rows:
        date_sheet.append(item)
    style_table(date_sheet)

    note = out.create_sheet("计算说明")
    note.append(["项目", "说明"])
    note.append(["数据来源", str(config.source)])
    note.append(["统计年月", f"{config.year}-{config.month:02d}"])
    note.append(["加班起算", config.overtime_start.strftime("%H:%M")])
    note.append(["加班规则", "每日最后一次有效打卡晚于加班起算时间的差额，向下按 30 分钟计入。"])
    note.append(["夜班识别", "存在 18:00 及以后打卡且无 10:00-14:30 白班/午间打卡，视为夜班开始。"])
    style_table(note)

    return out, len(rows), len(detail_rows)


def main() -> None:
    config = parse_args()
    try:
        workbook, employee_count, detail_count = build_report(config)
        config.output.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(config.output)
    except Exception as exc:
        raise SystemExit(f"生成失败: {exc}") from exc
    print(config.output)
    print(f"employees={employee_count} detail_rows={detail_count}")


if __name__ == "__main__":
    main()
